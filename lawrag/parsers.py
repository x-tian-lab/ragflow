# -*- coding: utf-8 -*-
"""Parser 层 (§4)。2026-07-13: XlsxParser 补回表注行过滤(M05/X03)。"""
import os
import re
from .schema import DocElement, Location
from . import law_regex as L

INVIS = re.compile(r'[​‌‍﻿　]')


def _clean(s: str) -> str:
    return INVIS.sub('', s).strip()


def law_elements(paras, doc_title):
    skip = L.strip_toc(paras)
    els = []
    cur_ch = cur_jie = None
    seq = 0
    for i, t in enumerate(paras):
        if i in skip:
            continue
        seq += 1
        t_sp = re.sub(r'\s+', ' ', t)
        base = [doc_title] if doc_title else []
        if L.RE_CH.match(t):
            cur_ch, cur_jie = t_sp, None
            els.append(DocElement('heading', t_sp, Location(heading_path=base + [t_sp], paragraph_seq=seq)))
        elif L.RE_JIE.match(t):
            cur_jie = t_sp
            els.append(DocElement('heading', t_sp, Location(
                heading_path=base + [x for x in (cur_ch, t_sp) if x], paragraph_seq=seq)))
        else:
            m = L.RE_TIAO.match(t)
            if m:
                els.append(DocElement('article', t_sp, Location(
                    heading_path=base + [x for x in (cur_ch, cur_jie) if x],
                    article_no=m.group(1), paragraph_seq=seq)))
            else:
                els.append(DocElement('paragraph', t_sp, Location(
                    heading_path=base + [x for x in (cur_ch, cur_jie) if x],
                    paragraph_seq=seq, anchor_text=t_sp[:20])))
    return els


class DocxParser:
    def parse(self, path, doc_title=''):
        from docx import Document
        d = Document(path)
        paras = []
        tables_by_pos = []
        ti = 0
        for child in d.element.body.iterchildren():
            tag = child.tag.rsplit('}', 1)[-1]
            if tag == 'p':
                txt = _clean(''.join(node.text or '' for node in child.iter()
                                     if node.tag.endswith('}t')))
                if txt:
                    paras.append(txt)
            elif tag == 'tbl':
                tables_by_pos.append((len(paras), d.tables[ti]))
                ti += 1
        if L.is_law_structured(paras):
            els = law_elements(paras, doc_title)
        else:
            els = [DocElement('paragraph', t, Location(paragraph_seq=s, anchor_text=t[:20]))
                   for s, t in enumerate(paras, 1)]
        for pos, tb in tables_by_pos:
            hdr = [_clean(c.text) for c in tb.rows[0].cells] if tb.rows else []
            for ri, row in enumerate(tb.rows[1:], 1):
                cells = [_clean(c.text) for c in row.cells]
                text = ' | '.join(('%s=%s' % (h, v)) if h else v for h, v in zip(hdr, cells) if v)
                if text:
                    els.append(DocElement('table_row', text, Location(
                        table_index=1, row_range='r%d' % ri, column_names=hdr or None,
                        paragraph_seq=pos)))
        return els


class TxtParser:
    HEADER_KEYS = ('标题：', '栏目：', '发布日期：', '来源：', '原文链接：')

    def parse(self, path, doc_title=''):
        raw = open(path, encoding='utf-8', errors='replace').read()
        m = re.search(r'^正文：\s*$', raw, re.M)
        body = raw[m.end():] if m else raw
        paras = [_clean(p) for p in body.split('\n')]
        paras = [p for p in paras if p and not p.startswith(self.HEADER_KEYS)
                 and p not in ('无障碍浏览',) and '下载文字版' not in p and '下载图片版' not in p]
        if L.is_law_structured(paras):
            return law_elements(paras, doc_title)
        return [DocElement('paragraph', t, Location(paragraph_seq=i, anchor_text=t[:20]))
                for i, t in enumerate(paras, 1)]


class MdParser:
    def parse(self, path, doc_title=''):
        lines = open(path, encoding='utf-8').read().split('\n')
        fm = {}
        if lines and lines[0].strip() == '---':
            try:
                end = lines[1:].index('---') + 1
                for ln in lines[1:end]:
                    if ':' in ln:
                        k, v = ln.split(':', 1)
                        fm[k.strip()] = v.strip()
                lines = lines[end + 1:]
            except ValueError:
                pass
        els = []
        hstack = []
        seq = 0
        table_hdr = None
        ti = 0
        for ln in lines:
            t = ln.rstrip()
            if not t.strip():
                continue
            seq += 1
            hm = re.match(r'^(#{1,4})\s+(.*)', t)
            if hm:
                lvl = len(hm.group(1))
                hstack = [(l, h) for l, h in hstack if l < lvl] + [(lvl, hm.group(2).strip())]
                els.append(DocElement('heading', hm.group(2).strip(),
                                      Location(heading_path=[h for _, h in hstack], paragraph_seq=seq)))
                table_hdr = None
                continue
            hp = [h for _, h in hstack]
            if t.startswith('|'):
                cells = [c.strip() for c in t.strip('|').split('|')]
                if all(re.fullmatch(r'-{3,}', c) for c in cells):
                    continue
                if table_hdr is None:
                    table_hdr = cells
                    ti += 1
                    continue
                text = ' | '.join('%s=%s' % (h, v) for h, v in zip(table_hdr, cells) if v)
                els.append(DocElement('table_row', text, Location(
                    heading_path=hp, sheet_name=fm.get('sheet_name'),
                    table_index=ti, row_range='r%d' % seq, column_names=table_hdr,
                    column_names_confidence=fm.get('column_names_confidence', 'ok'),
                    paragraph_seq=seq)))
                continue
            table_hdr = None
            body = re.sub(r'^[-*>]\s+', '', t.strip())
            m = L.RE_TIAO.match(body)
            els.append(DocElement('article' if m else 'paragraph', body, Location(
                heading_path=hp, article_no=m.group(1) if m else None,
                paragraph_seq=seq, anchor_text=body[:20])))
        return els


class XlsxParser:
    """R4/R5 + 表注行过滤(2026-07-13,M05/X03 检索缺口主因)。"""

    def parse(self, path, doc_title=''):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        els = []
        for ws in wb.worksheets:
            els.extend(self._sheet(ws))
        wb.close()
        return els

    @staticmethod
    def _cell_type(v):
        if v is None or not str(v).strip():
            return '_'
        if isinstance(v, (int, float)):
            return 'N'
        return 'N' if re.match(r'^-?[\d,，.%\s]+$', str(v).strip()) else 'T'

    def _sheet(self, ws):
        from collections import Counter
        for rng in list(ws.merged_cells.ranges):
            v = ws.cell(rng.min_row, rng.min_col).value
            ws.unmerge_cells(str(rng))
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    ws.cell(r, c).value = v
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            return []
        keep = [c for c in range(len(rows[0]))
                if any(r[c] is not None and str(r[c]).strip() for r in rows)]
        rows = [[r[c] for c in keep] for r in rows]
        sigs = [''.join(self._cell_type(v) for v in r) for r in rows]
        nonempty = [s for s in sigs if set(s) != {'_'}]
        if not nonempty:
            return []
        data_sig = Counter(nonempty).most_common(1)[0][0]
        data_start = sigs.index(data_sig)
        header_rows = []
        for i in range(data_start):
            if set(sigs[i]) == {'_'}:
                continue
            vals = [v for v in rows[i] if v is not None and str(v).strip()]
            if sum(1 for v in vals if self._cell_type(v) == 'N') / len(vals) > 0.3:
                continue
            if len(set(str(v).strip() for v in vals)) > 2:
                header_rows.append(i)
        headers = []
        for c in range(len(rows[0])):
            parts = []
            for r in header_rows:
                s = re.sub(r'\s+', ' ', str(rows[r][c]).strip()) if rows[r][c] is not None else ''
                if s and s not in parts:
                    parts.append(s)
            headers.append('.'.join(parts))
        conf = 'ok' if header_rows and len(header_rows) <= 3 and \
            sum(1 for h in headers if not h) < len(headers) * 0.3 else 'uncertain'
        els, ti, streak, last_note, last_note_ri = [], 1, 0, None, None
        for ri, r in enumerate(rows[data_start:], data_start + 1):
            vals = ['' if v is None else str(v).strip() for v in r]
            if not any(vals):
                streak += 1
                if streak == 2:
                    ti += 1
                continue
            streak = 0
            nums = [v for v in vals if re.fullmatch(r'\d{1,3}', v)]
            ne = [v for v in vals if v]
            if len(nums) >= 3 and len(nums) >= len(ne) - 1 and \
                    [int(v) for v in nums] == list(range(int(nums[0]), int(nums[0]) + len(nums))):
                continue
            sig = ''.join(self._cell_type(v) for v in r)
            # 表注判据:签名偏离数据签名,且 (宽表中非空格<=2 | 整行同值重复 | 长文本)
            # ——宽表(>=4列)的数据行不可能只有1-2个非空格;单细胞脚注由此捕获
            if sig != data_sig and len(set(ne)) <= 2 and (
                    (len(vals) >= 4 and len(ne) <= 2) or len(ne) >= 3
                    or max(len(v) for v in ne) > 30):
                # 长脚注常因合并单元格跨多行渲染,解并后逐行重复取值——仅紧邻行合并为一条(2026-07-14)
                if ne[0] == last_note and ri == last_note_ri + 1:
                    last_note_ri = ri
                    continue
                els.append(DocElement('paragraph', ne[0], Location(
                    sheet_name=ws.title, table_index=ti, paragraph_seq=ri,
                    anchor_text=ne[0][:20])))
                last_note, last_note_ri = ne[0], ri
                continue
            last_note, last_note_ri = None, None   # 普通数据行打断脚注连续性
            text = ' | '.join(('%s=%s' % (h, v)) if h else v for h, v in zip(headers, vals) if v)
            els.append(DocElement('table_row', text, Location(
                sheet_name=ws.title, table_index=ti, row_range='r%d' % ri,
                column_names=[h for h in headers if h] or None,
                column_names_confidence=conf)))
        return els


PARSERS = {'docx': DocxParser, 'doc': DocxParser, 'doc_converted': DocxParser,
           'txt': TxtParser, 'md': MdParser, 'xlsx': XlsxParser}


def route(file_path, fmt=None):
    fmt = fmt or os.path.splitext(file_path)[1].lstrip('.').lower()
    cls = PARSERS.get(fmt)
    if cls is None:
        raise ValueError('不支持的格式: %s (%s)' % (fmt, file_path))
    return cls()
